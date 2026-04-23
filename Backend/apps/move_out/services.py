from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from django.utils import timezone
import re
import os

import requests
from django.conf import settings
from openpyxl import load_workbook

from .models import MoveOutPlan

POSTCODES_API_BASE_URL = getattr(
    settings, "POSTCODES_API_BASE_URL", "https://api.postcodes.io"
)
ONS_PIPR_XLSX_URL = getattr(
    settings,
    "ONS_PIPR_XLSX_URL",
    (
        "https://www.ons.gov.uk/file?uri=%2Feconomy%2Finflationandpriceindices"
        "%2Fdatasets%2Fpriceindexofprivaterentsukmonthlypricestatistics"
        "%2F25march2026%2Fpriceindexofprivaterentsukmonthlypricestatistics.xlsx"
    ),
)
POLICE_API_BASE_URL = "https://data.police.uk/api"
APIFY_API_BASE_URL = "https://api.apify.com/v2"

APIFY_API_TOKEN = os.getenv("APIFY_API_TOKEN")

TIMEOUT = 15
MONEY = Decimal("0.01")

MONTH_NAMES = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


class MoveOutServiceError(Exception):
    pass


def lookup_postcode(postcode: str) -> dict:
    postcode = postcode.strip().upper()
    if not postcode:
        raise MoveOutServiceError("Postcode is needed")

    try:
        resp = requests.get(
            f"{POSTCODES_API_BASE_URL}/postcodes/{postcode}", timeout=TIMEOUT
        )
        resp.raise_for_status()
        result = resp.json().get("result")
    except requests.RequestException as e:
        raise MoveOutServiceError("Failed to look up postcode") from e

    if not result:
        raise MoveOutServiceError("Postcode is invalid")

    return {
        "postcode": result.get("postcode"),
        "country": result.get("country"),
        "region": result.get("region"),
        "admin_district": result.get("admin_district"),
        "area_code": result.get("codes", {}).get("admin_district"),
        "latitude": result.get("latitude"),
        "longitude": result.get("longitude"),
    }


def _to_text(value) -> str:
    return "" if value is None else str(value).strip().lower()


def _parse_date_from_row(values) -> date | None:
    for v in values:
        if isinstance(v, datetime):
            return date(v.year, v.month, 1)
        if isinstance(v, date):
            return date(v.year, v.month, 1)
        if isinstance(v, str):
            match = re.search(
                (
                    r"(january|february|march|april|may|june|july|august|"
                    r"september|october|november|december)\s+(\d{4})"
                ),
                v.strip().lower(),
            )
            if match:
                return date(int(match.group(2)), MONTH_NAMES[match.group(1)], 1)

    year = month = None
    for v in values:
        if isinstance(v, int) and 2015 <= v <= 2100:
            year = v
        if isinstance(v, str) and v.strip().lower() in MONTH_NAMES:
            month = MONTH_NAMES[v.strip().lower()]
        if isinstance(v, int) and 1 <= v <= 12 and month is None:
            month = v

    return date(year, month, 1) if year and month else None


def _find_rent_values(values) -> list[Decimal]:
    results = []
    for v in values:
        if v is None:
            continue
        if isinstance(v, int) and 2015 <= v <= 2100:
            continue
        try:
            n = Decimal(str(v))
        except (InvalidOperation, TypeError):
            continue
        if Decimal("100") <= n <= Decimal("10000"):
            results.append(n)
    return results


def _get_geography_name(postcode_data: dict) -> str:
    country = postcode_data.get("country")
    region = postcode_data.get("region")
    if country == "England" and region:
        return region
    if country:
        return country
    raise MoveOutServiceError("Could not determine geography for rent lookup")


def _download_ons_workbook():
    try:
        resp = requests.get(ONS_PIPR_XLSX_URL, timeout=TIMEOUT)
        resp.raise_for_status()
    except requests.RequestException as e:
        raise MoveOutServiceError("Failed to fetch ONS rent dataset") from e
    try:
        return load_workbook(
            filename=BytesIO(resp.content), read_only=True, data_only=True
        )
    except Exception as e:
        raise MoveOutServiceError("Could not open ONS rent dataset") from e


def _find_rent_for_area(workbook, area_name: str) -> Decimal | None:
    target = area_name.strip().lower()
    best_date = None
    best_rent = None

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row = list(row)
            if target not in [_to_text(v) for v in row]:
                continue

            row_date = _parse_date_from_row(row)
            rent_options = _find_rent_values(row)
            if not rent_options:
                continue

            rent = max(rent_options)
            if best_date is None and row_date is None and best_rent is None:
                best_rent = rent
            elif row_date and (best_date is None or row_date > best_date):
                best_date = row_date
                best_rent = rent

    return best_rent


def get_average_rent_from_ons(postcode_data: dict) -> Decimal:
    workbook = _download_ons_workbook()
    area = _get_geography_name(postcode_data)
    rent = _find_rent_for_area(workbook, area)

    if rent is None and postcode_data.get("country"):
        rent = _find_rent_for_area(workbook, postcode_data["country"])

    if rent is None:
        raise MoveOutServiceError(
            "Could not find rent data in ONS dataset for this area"
        )

    return rent.quantize(MONEY, rounding=ROUND_HALF_UP)


def get_crime_level(postcode_data: dict) -> str:
    lat = postcode_data.get("latitude")
    lng = postcode_data.get("longitude")
    if lat is None or lng is None:
        return "Unknown"

    try:
        resp = requests.get(
            f"{POLICE_API_BASE_URL}/crimes-street/all-crime",
            params={"lat": lat, "lng": lng},
            timeout=TIMEOUT,
        )
        resp.raise_for_status()
        count = len(resp.json())
    except (requests.RequestException, ValueError):
        return "Unknown"

    if count < 25:
        return "Very Low"
    if count < 60:
        return "Low"
    if count < 110:
        return "Moderate"
    if count < 180:
        return "High"
    return "Very High"


def get_rental_listings(search_location: str, max_price: Decimal) -> list[dict]:
    if not APIFY_API_TOKEN or not search_location:
        return []

    outcode = search_location.split()[0].upper()
    list_url = f"https://www.rightmove.co.uk/property-to-rent/{outcode}.html"

    run_input = {
        "listUrls": [{"url": list_url}],
        "fullPropertyDetails": False,
        "includePriceHistory": False,
        "includeNearestSchools": False,
        "maxProperties": 6,
        "proxy": {"useApifyProxy": True},
    }

    try:
        resp = requests.post(
            (
                f"{APIFY_API_BASE_URL}/acts/"
                "dhrumil~rightmove-scraper/run-sync-get-dataset-items"
            ),
            params={"token": APIFY_API_TOKEN},
            json=run_input,
            timeout=180,
        )
        print("APIFY STATUS:", resp.status_code)
        print("APIFY URL:", list_url)
        print("APIFY TEXT:", resp.text[:1000])
        resp.raise_for_status()
        items = resp.json()
    except Exception as e:
        print("APIFY ERROR:", repr(e))
        return []

    listings = []
    for item in items:
        price = item.get("price")
        if price is not None:
            try:
                numeric_price = Decimal(re.sub(r"[^\d.]", "", str(price)))
                if numeric_price > max_price:
                    continue
            except Exception:
                pass

        images = item.get("images") or []
        image_url = images[0] if images else None

        listings.append(
            {
                "listing_id": item.get("id") or item.get("propertyId"),
                "display_address": item.get("displayAddress") or item.get("address"),
                "latest_price": price,
                "display_price": item.get("displayPrice"),
                "bedrooms": item.get("bedrooms"),
                "bathrooms": item.get("bathrooms"),
                "property_type": item.get("propertyType"),
                "property_sub_type": item.get("propertySubType"),
                "agent": item.get("agent"),
                "agent_branch": item.get("agentBranch"),
                "added_date": item.get("addedOn"),
                "image_url": image_url,
                "listing_url": item.get("url"),
                "source": "Rightmove via Apify",
            }
        )

        if len(listings) == 6:
            break

    print("FINAL LISTINGS:", listings)
    return listings


def calculate_disposable_income(income: Decimal, expenses: Decimal) -> Decimal:
    return (income - expenses).quantize(MONEY, rounding=ROUND_HALF_UP)


def calculate_rent_ratio_percent(income: Decimal, rent: Decimal) -> Decimal:
    if income <= 0:
        return Decimal("0.00")
    return (rent / income * 100).quantize(MONEY, rounding=ROUND_HALF_UP)


def calculate_readiness(income: Decimal, expenses: Decimal, rent: Decimal) -> dict:
    disposable = calculate_disposable_income(income, expenses)
    rent_ratio = calculate_rent_ratio_percent(income, rent)

    if disposable <= 0:
        status, score = MoveOutPlan.ReadinessStatus.not_ready, 10
    elif disposable < rent:
        status, score = MoveOutPlan.ReadinessStatus.needs_improvement, 35
    elif rent_ratio <= Decimal("30.00"):
        status, score = MoveOutPlan.ReadinessStatus.ready, 85
    else:
        status, score = MoveOutPlan.ReadinessStatus.borderline, 60

    return {
        "disposable_income": disposable,
        "rent_ratio_percent": rent_ratio,
        "status": status,
        "readiness_score": score,
    }


def generate_summary(status: str, estimated_rent: Decimal, crime_level: str) -> str:
    crime_text = f" Crime level in the area appears to be {crime_level.lower()}."

    messages = {
        MoveOutPlan.ReadinessStatus.ready: (
            "You appear ready to move out. Your finances look strong compared "
            f"with the estimated local rent of £{estimated_rent}.{crime_text}"
        ),
        MoveOutPlan.ReadinessStatus.borderline: (
            "You may be able to move out, but affordability looks tight compared "
            f"with the estimated local rent of £{estimated_rent}.{crime_text}"
        ),
        MoveOutPlan.ReadinessStatus.needs_improvement: (
            "Your remaining monthly income is currently below the estimated rent "
            f"of £{estimated_rent}, so moving out may be difficult right now."
            f"{crime_text}"
        ),
    }

    return messages.get(
        status,
        (
            "You are not currently in a strong position to move out based on "
            "your income, expenses, and the rent estimate for your area."
            f"{crime_text}"
        ),
    )


def save_move_out_plan(
    user, postcode: str, monthly_income, monthly_expenses
) -> MoveOutPlan:
    try:
        monthly_income = Decimal(str(monthly_income))
        monthly_expenses = Decimal(str(monthly_expenses))
    except (InvalidOperation, TypeError) as e:
        raise MoveOutServiceError("Income and expenses must be valid numbers") from e

    postcode_data = lookup_postcode(postcode)
    estimated_rent = get_average_rent_from_ons(postcode_data)
    crime_level = get_crime_level(postcode_data)
    readiness = calculate_readiness(monthly_income, monthly_expenses, estimated_rent)

    existing = get_saved_move_out_plan(user)
    postcode_changed = (
        not existing or existing.target_postcode != postcode_data["postcode"]
    )
    stale = not existing or (timezone.now() - existing.updated_at) > timedelta(days=3)

    if postcode_changed or stale:
        search_location = (
            postcode_data.get("postcode")
            or postcode_data.get("admin_district")
            or postcode_data.get("region")
            or ""
        )
        listings = get_rental_listings(search_location, readiness["disposable_income"])
    else:
        listings = existing.property_listings

    summary = generate_summary(readiness["status"], estimated_rent, crime_level)

    plan, _ = MoveOutPlan.objects.update_or_create(
        user=user,
        defaults={
            "target_postcode": postcode_data["postcode"],
            "area_name": postcode_data.get("admin_district")
            or postcode_data.get("region")
            or postcode_data.get("country")
            or "",
            "area_code": postcode_data.get("area_code") or "",
            "monthly_income": monthly_income,
            "monthly_expenses": monthly_expenses,
            "estimated_monthly_rent": estimated_rent,
            "disposable_income": readiness["disposable_income"],
            "rent_ratio_percent": readiness["rent_ratio_percent"],
            "readiness_score": readiness["readiness_score"],
            "status": readiness["status"],
            "crime_level": crime_level,
            "property_listings": listings,
            "summary": summary,
        },
    )

    return plan


def get_saved_move_out_plan(user):
    try:
        return user.move_out_plan
    except MoveOutPlan.DoesNotExist:
        return None
